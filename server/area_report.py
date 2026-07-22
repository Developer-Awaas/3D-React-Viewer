"""Render a computed area statement into a professional .xlsx a developer can
hand over. Pure function (bytes in memory) so every input case is unit-tested
before it is wired to an endpoint.

Input is the dict from area_statement.compute_area_statement(). This module does
NO geometry — it only formats, so a bad statement can never crash a parse.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_HDR_FILL = PatternFill("solid", fgColor="1F2A44")     # deep navy
_HDR_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=16, color="1F2A44")
_BOLD = Font(bold=True)
_MUTED = Font(color="6B7280", size=9)
_THIN = Side(style="thin", color="D0D5DD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RIGHT = Alignment(horizontal="right")


def _num(v, default=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def build_area_xlsx(statement, project="", plan_name="", generated_on=""):
    """area statement dict -> .xlsx bytes. Defensive: missing keys default to 0/blank."""
    statement = statement or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Area Statement"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    r = 1
    ws.cell(r, 1, "RERA Area Statement").font = _TITLE_FONT
    r += 1
    if project:
        ws.cell(r, 1, f"Project: {project}"); r += 1
    if plan_name:
        ws.cell(r, 1, f"Plan: {plan_name}"); r += 1
    if generated_on:
        c = ws.cell(r, 1, f"Generated: {generated_on}"); c.font = _MUTED; r += 1
    r += 1

    # per-room carpet table
    for j, h in enumerate(("Room", "Carpet (sq ft)", "Carpet (sq m)"), start=1):
        c = ws.cell(r, j, h)
        c.fill = _HDR_FILL; c.font = _HDR_FONT; c.border = _BORDER
    r += 1
    for room in statement.get("rooms", []) or []:
        ws.cell(r, 1, str(room.get("id", ""))).border = _BORDER
        c = ws.cell(r, 2, round(_num(room.get("sqft")), 1)); c.border = _BORDER; c.alignment = _RIGHT
        c = ws.cell(r, 3, round(_num(room.get("sqm")), 2)); c.border = _BORDER; c.alignment = _RIGHT
        r += 1

    # summary block
    def row(label, pair_or_val, unit_pair=True, bold=False):
        nonlocal r
        lc = ws.cell(r, 1, label)
        lc.border = _BORDER
        if bold:
            lc.font = _BOLD
        if unit_pair:
            p = pair_or_val or {}
            c = ws.cell(r, 2, round(_num(p.get("sqft")), 1)); c.border = _BORDER; c.alignment = _RIGHT
            c = ws.cell(r, 3, round(_num(p.get("sqm")), 2)); c.border = _BORDER; c.alignment = _RIGHT
            if bold:
                ws.cell(r, 2).font = _BOLD; ws.cell(r, 3).font = _BOLD
        else:
            c = ws.cell(r, 2, pair_or_val); c.border = _BORDER; c.alignment = _RIGHT
            ws.cell(r, 3).border = _BORDER
        r += 1

    r += 1
    row("Total carpet area", statement.get("carpet_area"), bold=True)
    row("Built-up area", statement.get("built_up_area"))
    row("Super built-up area", statement.get("super_built_up_area"), bold=True)
    row("Wall & circulation", statement.get("wall_and_circulation"))
    row("Loading factor", _num(statement.get("loading_factor"), 1.30), unit_pair=False)
    row("Carpet efficiency (%)", _num(statement.get("efficiency_pct")), unit_pair=False)

    # notes + disclaimer
    r += 1
    for note in statement.get("notes", []) or []:
        c = ws.cell(r, 1, f"Note: {note}"); c.font = _MUTED
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1
    disc = statement.get("disclaimer")
    if disc:
        c = ws.cell(r, 1, disc); c.font = Font(italic=True, color="B45309")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
