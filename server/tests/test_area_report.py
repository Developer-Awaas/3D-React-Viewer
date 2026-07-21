"""Tests for the .xlsx area-statement builder. Each input case is checked
individually BEFORE the endpoint is exercised: a full statement, empty rooms,
missing keys, and bad value types must all produce a valid workbook."""
import io

import openpyxl

import area_report

FULL = {
    "carpet_area": {"sqft": 200.0, "sqm": 18.58},
    "built_up_area": {"sqft": 300.0, "sqm": 27.87},
    "super_built_up_area": {"sqft": 390.0, "sqm": 36.23},
    "wall_and_circulation": {"sqft": 100.0, "sqm": 9.29},
    "loading_factor": 1.30,
    "efficiency_pct": 51.3,
    "rooms": [{"id": "r0", "sqft": 120.0, "sqm": 11.15},
              {"id": "r1", "sqft": 80.0, "sqm": 7.43}],
    "notes": ["carpet excludes internal walls"],
    "disclaimer": "Estimate — verify with an architect.",
}


def _load(b):
    assert b[:2] == b"PK"                      # xlsx is a zip
    return openpyxl.load_workbook(io.BytesIO(b)).active


def test_full_statement_builds_valid_xlsx():
    ws = _load(area_report.build_area_xlsx(FULL, project="Awaas Heights",
                                           plan_name="Unit A", generated_on="2026-07-21"))
    flat = [ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1) for c in range(1, 4)]
    assert "RERA Area Statement" in flat
    assert "Project: Awaas Heights" in flat
    assert 200.0 in flat        # total carpet
    assert 390.0 in flat        # super built-up
    assert "r0" in flat and "r1" in flat


def test_empty_rooms_still_builds():
    st = {**FULL, "rooms": []}
    ws = _load(area_report.build_area_xlsx(st))
    assert ws.max_row >= 5      # header + summary still present


def test_missing_keys_default_safely():
    ws = _load(area_report.build_area_xlsx({}))     # nothing at all
    assert ws["A1"].value == "RERA Area Statement"  # never crashes


def test_bad_value_types_do_not_crash():
    st = {"carpet_area": {"sqft": "oops", "sqm": None},
          "rooms": [{"id": 5, "sqft": None, "sqm": "x"}],
          "loading_factor": "n/a"}
    b = area_report.build_area_xlsx(st)
    assert b[:2] == b"PK"       # still a valid workbook (bad numbers -> 0)


def test_returns_bytes_not_none():
    b = area_report.build_area_xlsx(FULL)
    assert isinstance(b, (bytes, bytearray)) and len(b) > 500


def test_endpoint_returns_xlsx():
    import os

    from fastapi.testclient import TestClient

    import main
    fixture = os.path.join(os.path.dirname(__file__), "fixtures", "plan_20x45_flat.pdf")
    if not os.path.exists(fixture):
        return
    c = TestClient(main.app)
    with open(fixture, "rb") as f:
        r = c.post("/area-statement.xlsx?width_ft=20&project=Awaas",
                   files={"image": ("20x45.pdf", f.read(), "application/pdf")})
    assert r.status_code == 200
    assert r.content[:2] == b"PK"
    assert "spreadsheet" in r.headers["content-type"]
